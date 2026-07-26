"""
ZeroAnomalías — VAE Frontend
Streamlit app con sistema de login y dos perfiles: Técnico y Negocio.
"""
# pylint: disable=line-too-long
# Varias líneas largas son bloques de HTML/CSS dentro de f-strings para las
# tarjetas KPI y el tema oscuro. Partirlas a la mitad las volvería HTML
# inválido o forzaría concatenaciones que complican más de lo que ayudan.
#
# pylint: disable=use-dict-literal
# Los dict(...) de Plotly (go.Figure, layout, marker, etc.) se dejan como
# llamada a dict() en vez de {...} literal a propósito: es el estilo que usa
# la documentación oficial de Plotly y así queda igual en todo el archivo.
# No cambia el comportamiento, solo el estilo de escritura.
#
# pylint: disable=invalid-name
# Este archivo es un script de Streamlit: la mayoría de "constantes" que
# Pylint detecta (ej. umbrales, dataframes, figuras) en realidad son
# variables normales dentro del flujo secuencial de la página, no
# constantes de módulo. Forzar MAYUSCULAS en todas rompería la legibilidad
# del código sin aportar nada.
#
# pylint: disable=too-many-lines
# El archivo agrupa las 3 secciones del perfil Técnico y el perfil de
# Negocio en un solo script de Streamlit (así arrancó el proyecto). Partirlo
# en módulos es una mejora válida a futuro, pero implica reestructurar el
# manejo de session_state y no se hace a días de la sustentación para no
# arriesgar nada que ya funciona y está probado.
import json
import io
import re
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.logger import log_info, log_warning, LOG_FILE
from src.audit_service import process_batch, evaluate_raw_transaction, process_raw_batch

# ── Configuración de página ──────────────────────────────────────────────────
st.set_page_config(
    page_title="ZeroAnomalías",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded",
)

LOGO_FULL = "assets/logo_zeroanomalias.jpeg"
LOGO_ICON = "assets/logo_icon_only.png"

# ── Paleta de colores de severidad ───────────────────────────────────────────
SEVERITY_COLORS = {
    "normal": "#22c55e",
    "baja":   "#eab308",
    "media":  "#f97316",
    "alta":   "#ef4444",
}
SEVERITY_ORDER = ["normal", "baja", "media", "alta"]
SEV_LABELS     = {"normal": "Normal", "baja": "Baja", "media": "Media", "alta": "Alta"}

# ── Traducción de tipos de anomalía reales (dataset de evaluación) a texto ──
TIPO_ANOMALIA_TECNICO = {
    "monto_extremo": "Monto fuera del rango estadístico normal",
    "devolucion_monto_alto": "Devolución de un monto inusualmente alto",
    "descuento_excesivo": "Descuento fuera del rango permitido (>30%)",
    "hora_inusual": "Transacción fuera del horario de operación normal",
    "anulacion_sospechosa": "Anulación con patrón atípico",
}
TIPO_ANOMALIA_NEGOCIO = {
    "monto_extremo": "Monto mucho más alto o más bajo de lo normal",
    "devolucion_monto_alto": "Devolución grande — vale la pena revisarla con el cajero",
    "descuento_excesivo": "Descuento más grande de lo que se suele permitir",
    "hora_inusual": "Venta hecha en un horario poco común",
    "anulacion_sospechosa": "Anulación con un patrón que no es el habitual",
}


def explicar_anomalia(row, estilo="tecnico"):  # pylint: disable=too-many-branches
    # too-many-branches: cada rama es una regla de negocio independiente
    # (tipo de monto, descuento, horario); dividirlas en sub-funciones no
    # las simplifica, solo las esparce. Los nombres locales (severidad,
    # monto, hora, factores) coinciden con variables de página usadas más
    # abajo en "Verificar Venta" — no hay ambigüedad real, se marca inline
    # donde corresponde.
    """
    Genera una explicación corta y legible de por qué una transacción se
    marcó como anómala, usando los campos que estén disponibles en `row`
    (funciona tanto con el dataset de evaluación del Técnico —que trae la
    categoría real `tipo_anomalia`— como con archivos nuevos sin esa
    etiqueta, como los que sube Rosita, infiriendo la razón con reglas).

    `estilo`: "tecnico" (más cuantitativo) o "negocio" (lenguaje simple).
    Devuelve "" si la transacción es normal.
    """
    severidad = str(row.get("severidad", row.get("estado", ""))).lower()  # pylint: disable=redefined-outer-name
    if severidad in ("", "normal", "nan"):
        return ""

    # 1) Explicación basada en el modelo real: qué variable(s) concentraron
    #    más error de reconstrucción para ESTA transacción específica
    #    (calculada en src/audit_service.py). Es la más precisa porque sale
    #    directo del VAE, no de reglas a mano, y sí detecta combinaciones
    #    de factores (ej. "hora + mesa" juntos), no solo uno a la vez.
    motivo_col = "motivo_tecnico" if estilo == "tecnico" else "motivo_negocio"
    motivo_desde_modelo = row.get(motivo_col)
    if isinstance(motivo_desde_modelo, str) and motivo_desde_modelo.strip():
        return motivo_desde_modelo

    # 2) Si existe la categoría real (solo en el dataset de evaluación), es
    #    la explicación más precisa posible: el motivo con el que esa
    #    transacción fue generada.
    tipo_real = row.get("tipo_anomalia")
    if isinstance(tipo_real, str) and tipo_real.lower() not in ("", "normal", "nan"):
        etiquetas = TIPO_ANOMALIA_TECNICO if estilo == "tecnico" else TIPO_ANOMALIA_NEGOCIO
        return etiquetas.get(tipo_real.lower(), tipo_real.replace("_", " ").capitalize())

    # 3) Sin lo anterior: inferir con reglas simples sobre los campos crudos
    #    disponibles (respaldo para cuando el error está muy repartido y el
    #    modelo no señala una variable dominante clara).
    monto = row.get("monto", row.get("monto_final"))  # pylint: disable=redefined-outer-name
    descuento_monto = row.get("descuento")
    descuento_pct = row.get("descuento_pct")
    if descuento_monto is None and descuento_pct is not None and monto is not None:
        try:
            descuento_monto = float(monto) * float(descuento_pct)
        except (TypeError, ValueError):
            descuento_monto = None

    hora = row.get("hora")  # pylint: disable=redefined-outer-name
    if hora is None and row.get("fecha_hora") is not None:
        try:
            hora = pd.to_datetime(row["fecha_hora"]).hour
        except (ValueError, TypeError):
            hora = None

    factores = []  # pylint: disable=redefined-outer-name
    try:
        monto_f = float(monto) if monto is not None else None
    except (TypeError, ValueError):
        monto_f = None

    if monto_f is not None and monto_f > 200:
        factores.append(
            f"Monto ${monto_f:,.2f} supera el umbral típico de $200" if estilo == "tecnico"
            else "Monto mucho más alto de lo habitual"
        )
    elif monto_f is not None and monto_f < 1.0:
        factores.append(
            f"Monto ${monto_f:,.2f} atípicamente bajo" if estilo == "tecnico"
            else "Monto inusualmente bajo"
        )
    if descuento_monto is not None and monto_f:
        ratio = float(descuento_monto) / max(monto_f, 1)
        if ratio > 0.3:
            factores.append(
                f"Descuento del {ratio*100:.0f}% supera el 30% permitido" if estilo == "tecnico"
                else "Descuento fuera de lo normal"
            )
    if hora is not None and (hora < 6 or hora > 22):
        factores.append(
            f"Hora inusual ({int(hora):02d}:00)" if estilo == "tecnico"
            else "Venta en un horario poco común"
        )

    if factores:
        return " · ".join(factores[:2])
    return (
        "Patrón general inusual detectado por el modelo VAE" if estilo == "tecnico"
        else "Patrón inusual detectado, sin una causa puntual clara"
    )


# ── Paleta para los reportes Excel descargables (mismos tonos que el dashboard) ──
EXCEL_FILL_HEX = {
    "normal": "DCFCE7",  # verde pastel
    "baja":   "FEF9C3",  # amarillo pastel
    "media":  "FFEDD5",  # naranja pastel
    "alta":   "FEE2E2",  # rojo pastel
}
EXCEL_FONT_HEX = {
    "normal": "166534",
    "baja":   "854D0E",
    "media":  "9A3412",
    "alta":   "991B1B",
}


LOG_LINE_RE = re.compile(
    r"^\[(?P<ts>[^\]]+)\]\s*\[(?P<level>[A-Z]+)\]\s*\[(?P<event>[A-Za-z_]+)\]\s*(?P<details>.*)$"
)
LOG_NUM_FIELD_RE = re.compile(r"(?P<key>\w+)=(?P<value>[\d.]+)")


def parse_log_lines(lines):
    """
    Convierte líneas crudas de logs/app.log (formato
    "[fecha] [NIVEL] [EVENTO] campo=valor, campo=valor") en un DataFrame
    con columnas: timestamp, level, event, latencia_ms, throughput.
    Líneas que no calzan con el formato esperado se ignoran.
    """
    rows = []
    for line in lines:
        match = LOG_LINE_RE.match(line)
        if not match:
            continue
        details = match.group("details")
        fields = {k: float(v) for k, v in LOG_NUM_FIELD_RE.findall(details)}
        rows.append({
            "timestamp": match.group("ts"),
            "level": match.group("level"),
            "event": match.group("event"),
            "latencia_ms": fields.get("latencia", None) * 1000
                           if "latencia" in fields else None,
            "throughput": fields.get("throughput", None),
        })
    return pd.DataFrame(rows, columns=["timestamp", "level", "event", "latencia_ms", "throughput"])


def build_styled_excel_bytes(sheets):  # pylint: disable=too-many-locals,too-many-statements
    # Función de formateo de Excel: recorre cada hoja, cada fila y cada
    # columna aplicando fuente/color/borde/formato, así que naturalmente usa
    # más variables y líneas de las que Pylint recomienda por defecto.
    # Partirla en sub-funciones más pequeñas no simplifica nada, solo mueve
    # el mismo trabajo a otro lugar.
    """
    Genera un Excel (.xlsx) descargable con encabezado oscuro, encabezado
    congelado, filtros automáticos, columnas autoajustadas y filas coloreadas
    por severidad — mismo lenguaje visual que el resto del dashboard.

    `sheets` es una lista de dicts, cada uno puede traer:
      - name (str, obligatorio): nombre de la pestaña (máx. 31 caracteres)
      - df (DataFrame, obligatorio): datos ya con encabezados amigables
      - sev_series (Series, opcional): severidad ("normal"/"baja"/"media"/
        "alta") alineada por posición con `df`, para colorear cada fila
      - currency_cols (list[str], opcional): columnas a formatear como $
      - decimal_cols (list[str], opcional): columnas a formatear con 6 decimales
      - highlight_col / highlight_values (opcional): resalta filas de una
        hoja resumen sin severidad (ej. la fila de "Posibles pérdidas")

    Devuelve un BytesIO listo para pasar a st.download_button.
    """
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_align = Alignment(horizontal="center", vertical="center")
    default_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet in sheets:
        df = sheet["df"]
        sev_series = sheet.get("sev_series")
        currency_cols = sheet.get("currency_cols", [])
        decimal_cols = sheet.get("decimal_cols", [])
        highlight_col = sheet.get("highlight_col")
        highlight_values = [v.lower() for v in sheet.get("highlight_values", [])]

        ws = wb.create_sheet(title=sheet["name"][:31])

        for c_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
        ws.row_dimensions[1].height = 22

        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            row_sev = None
            if sev_series is not None and (r_idx - 2) < len(sev_series):
                row_sev = str(sev_series.iloc[r_idx - 2]).lower()

            row_is_highlight = False
            if highlight_col and highlight_col in df.columns:
                cell_text = str(row[highlight_col]).lower()
                row_is_highlight = any(v in cell_text for v in highlight_values)

            for c_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=row[col_name])
                cell.border = border

                if row_sev in EXCEL_FILL_HEX:
                    cell.fill = PatternFill("solid", fgColor=EXCEL_FILL_HEX[row_sev])
                    cell.font = Font(name="Arial", size=10, color=EXCEL_FONT_HEX[row_sev])
                elif row_is_highlight:
                    cell.fill = PatternFill("solid", fgColor=EXCEL_FILL_HEX["alta"])
                    cell.font = Font(name="Arial", size=10, bold=True, color=EXCEL_FONT_HEX["alta"])
                else:
                    cell.font = default_font

                if col_name in currency_cols:
                    cell.number_format = '"$"#,##0.00'
                elif col_name in decimal_cols:
                    cell.number_format = "0.000000"

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for c_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(c_idx)
            values_len = [len(str(v)) for v in df[col_name].astype(str).tolist()]
            max_len = max([len(str(col_name))] + values_len) if values_len else len(str(col_name))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 42)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Credenciales hardcoded ────────────────────────────────────────────────────
USERS = {
    "tecnico": {"password": "admin123", "rol": "Técnico",  "display": "Técnico VAE"},
    "rosita":  {"password": "rosita123", "rol": "Negocio", "display": "Restaurante Rosita"},
}

# ── Inicializar session_state ─────────────────────────────────────────────────
if "logged_in" not in st.session_state:
    st.session_state.logged_in  = False
    st.session_state.username   = ""
    st.session_state.rol        = ""
    st.session_state.display    = ""
    st.session_state.login_err  = ""


# ── CSS global ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .stApp { background-color: #0f172a; color: #e2e8f0; }
    section[data-testid="stSidebar"] {
        background-color: #1e293b;
        border-right: 1px solid #334155;
    }
    .kpi-card {
        background: #1e293b;
        border: 1px solid #334155;
        border-radius: 12px;
        padding: 20px 24px;
        text-align: center;
    }
    .kpi-title { font-size:0.85rem; color:#94a3b8; text-transform:uppercase;
                 letter-spacing:0.08em; margin-bottom:6px; }
    .kpi-value { font-size:2rem; font-weight:700; color:#f1f5f9; }
    .kpi-sub   { font-size:0.78rem; color:#64748b; margin-top:4px; }
    h2, h3 { color:#f1f5f9 !important; }
    .stTabs [data-baseweb="tab-list"] { background-color:#1e293b; gap:4px; }
    .stTabs [data-baseweb="tab"] {
        background-color:#0f172a; color:#94a3b8;
        border-radius:8px 8px 0 0; padding:8px 20px;
        border:1px solid #334155;
    }
    .stTabs [aria-selected="true"] {
        background-color:#334155 !important; color:#f1f5f9 !important;
    }
    /* Login card */
    .login-card {
        max-width: 420px;
        margin: 60px auto;
        background: #1e293b;
        border: 1px solid #334155;
        border-radius: 16px;
        padding: 40px 36px 32px;
    }
    /* Rosita card */
    .rosita-welcome {
        background: linear-gradient(135deg, #1e293b 0%, #0f172a 100%);
        border: 1px solid #f97316;
        border-radius: 16px;
        padding: 28px 32px;
        margin-bottom: 24px;
    }
</style>
""", unsafe_allow_html=True)

# ── Layout de Plotly (tema oscuro) ────────────────────────────────────────────
PLOTLY_LAYOUT = dict(
    paper_bgcolor="#1e293b",
    plot_bgcolor="#1e293b",
    font=dict(color="#e2e8f0", family="Inter, sans-serif"),
    margin=dict(l=20, r=20, t=40, b=20),
)


# ════════════════════════════════════════════════════════════════════════════
# PANTALLA DE LOGIN
# ════════════════════════════════════════════════════════════════════════════
def show_login():
    """Renderiza la pantalla de inicio de sesión."""
    # Centrar el formulario con columnas
    _, center, _ = st.columns([1, 1.2, 1])
    with center:
        st.markdown('<div style="padding-top:24px;"></div>', unsafe_allow_html=True)
        st.image(LOGO_FULL, use_container_width=True)
        st.markdown("""
        <div style="text-align:center; padding: 0 0 20px;">
          <p style="color:#64748b; font-size:0.9rem;">
            Ingresa tus credenciales para continuar
          </p>
        </div>
        """, unsafe_allow_html=True)

        with st.form("login_form", clear_on_submit=False):
            usuario = st.text_input("👤 Usuario", placeholder="Ingresa tu usuario")
            clave   = st.text_input("🔒 Contraseña", type="password",
                                    placeholder="Ingresa tu contraseña")
            login_submitted = st.form_submit_button("Iniciar Sesión", use_container_width=True)

            if login_submitted:
                user_data = USERS.get(usuario.strip().lower())

                if user_data and clave == user_data["password"]:
                    st.session_state.logged_in = True
                    st.session_state.username = usuario.strip().lower()
                    st.session_state.rol = user_data["rol"]
                    st.session_state.display = user_data["display"]
                    st.session_state.login_err = ""

                    log_info(
                        "LOGIN_OK",
                        f"usuario={st.session_state.username}, rol={st.session_state.rol}"
                    )

                    st.rerun()

                else:
                    st.session_state.login_err = "Usuario o contraseña incorrectos."

                    log_warning(
                        "LOGIN_FAIL",
                        f"usuario={usuario}"
                    )

        if st.session_state.login_err:
            st.error(st.session_state.login_err)

        st.markdown("""
        <div style="text-align:center; margin-top:20px; color:#475569; font-size:0.78rem;">
          Demo — credenciales de prueba disponibles
        </div>
        """, unsafe_allow_html=True)


if not st.session_state.logged_in:
    show_login()
    st.stop()


# ── Carga de datos con caché (solo para rol Técnico) ─────────────────────────
@st.cache_data
def load_metricas():
    """Carga las métricas de evaluación del modelo (precision, recall, F1, etc.)."""
    with open("reports/metricas_evaluacion.json", encoding="utf-8") as f:
        return json.load(f)

@st.cache_data
def load_transacciones():
    """Carga el CSV de transacciones ya evaluadas por el modelo VAE."""
    return pd.read_csv("reports/evaluacion_transacciones.csv")

@st.cache_data
def load_training_history():
    """Carga el historial de pérdida (loss) del entrenamiento del VAE."""
    return pd.read_csv("reports/vae_training_history.csv")

@st.cache_data
def load_umbrales():
    """Carga los umbrales de severidad (baja/media/alta) calibrados en validación."""
    with open("reports/umbral_severidad.json", encoding="utf-8") as f:
        return json.load(f)


# ── Sidebar compartido (botón de logout) ─────────────────────────────────────
with st.sidebar:
    lc_a, lc_b, lc_c = st.columns([1, 2, 1])
    with lc_b:
        st.image(LOGO_ICON, use_container_width=True)
    st.markdown(
        """<div style="text-align:center; font-weight:800; color:#f1f5f9;
                    font-size:1.05rem; letter-spacing:0.02em; margin-top:-8px;">
            ZeroAnomalías
          </div>""",
        unsafe_allow_html=True,
    )

    rol_icon = "🛠️" if st.session_state.rol == "Técnico" else "🍽️"
    st.markdown(
        f"""
        <div style="padding:12px 0 8px;">
          <div style="font-size:1.6rem; text-align:center;">{rol_icon}</div>
          <div style="text-align:center; font-weight:700; color:#f1f5f9;
                      font-size:0.95rem; margin-top:4px;">
            {st.session_state.display}
          </div>
          <div style="text-align:center; color:#64748b; font-size:0.78rem; margin-top:2px;">
            Rol: {st.session_state.rol}
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown("---")

    if st.session_state.rol == "Técnico":
        st.markdown("#### Navegación")
        seccion = st.radio(
            "",
            ["📊 Metricas de Rendimiento", "📁 Auditoría por Lotes (CSV)", "⚡ Verificar Venta", "📜 Logs y Rendimiento Sistema"],
            label_visibility="collapsed",
        )
        umbrales = load_umbrales()
        st.markdown("---")
        st.markdown("#### Umbrales de Severidad")
        st.markdown(
            f"""
            <div style='font-size:0.82rem; line-height:1.9;'>
              <span style='color:{SEVERITY_COLORS["baja"]}'>●</span>
              <b>Baja</b> &nbsp;≥ {umbrales['umbral_baja']:.4f}<br>
              <span style='color:{SEVERITY_COLORS["media"]}'>●</span>
              <b>Media</b> &nbsp;≥ {umbrales['umbral_media']:.4f}<br>
              <span style='color:{SEVERITY_COLORS["alta"]}'>●</span>
              <b>Alta</b> &nbsp;≥ {umbrales['umbral_alta']:.4f}
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.markdown("---")
        st.caption("Modelo VAE · 50 épocas · Split: prueba")
    else:
        seccion = None  # No aplica para Negocio

    st.markdown("---")
    if st.button("🚪 Cerrar Sesión", use_container_width=True):
        for key in ["logged_in", "username", "rol", "display", "login_err"]:
            st.session_state[key] = False if key == "logged_in" else ""
        st.rerun()


# ════════════════════════════════════════════════════════════════════════════
# ENRUTAMIENTO POR ROL
# ════════════════════════════════════════════════════════════════════════════
if st.session_state.rol == "Técnico":

    metricas   = load_metricas()
    df_tx      = load_transacciones()
    df_history = load_training_history()
    umbrales   = load_umbrales()
    m          = metricas["prueba"]

    # ────────────────────────────────────────────────────────────────────────
    # SECCIÓN 1 — DASHBOARD GENERAL
    # ────────────────────────────────────────────────────────────────────────
    if seccion == "📊 Metricas de Rendimiento":
        st.markdown("## 📊 Metricas de Rendimiento")
        st.markdown("Resumen del desempeño del modelo VAE y análisis de anomalías detectadas.")

        k1, k2, k3, k4 = st.columns(4)
        with k1:
            st.markdown(
                f"""<div class="kpi-card">
                  <div class="kpi-title">Precisión (Prueba)</div>
                  <div class="kpi-value" style="color:#22c55e;">{m['precision']*100:.1f}%</div>
                  <div class="kpi-sub">F1-Score: {m['f1_score']*100:.1f}%</div>
                </div>""", unsafe_allow_html=True)
        with k2:
            st.markdown(
                f"""<div class="kpi-card">
                  <div class="kpi-title">Total Anomalías</div>
                  <div class="kpi-value" style="color:#f97316;">{m['n_predichas_anomalas']:,}</div>
                  <div class="kpi-sub">de {m['n_transacciones']:,} transacciones</div>
                </div>""", unsafe_allow_html=True)
        with k3:
            st.markdown(
                f"""<div class="kpi-card">
                  <div class="kpi-title">Monto en Riesgo</div>
                  <div class="kpi-value" style="color:#ef4444;">${m['monto_total_en_riesgo']:,.2f}</div>
                  <div class="kpi-sub">{m['pct_monto_en_riesgo']:.1f}% del total</div>
                </div>""", unsafe_allow_html=True)
        with k4:
            st.markdown(
                f"""<div class="kpi-card">
                  <div class="kpi-title">Recall</div>
                  <div class="kpi-value" style="color:#a78bfa;">{m['recall']*100:.1f}%</div>
                  <div class="kpi-sub">Anomalías reales detectadas</div>
                </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # Gráficos: pastel + training history
        col_pie, col_line = st.columns([1, 1.6])
        with col_pie:
            st.markdown("### Distribución de Severidad")
            sev_dist   = m["distribucion_severidad"]
            sev_labels = [SEV_LABELS[k] for k in SEVERITY_ORDER if k in sev_dist]
            sev_values = [sev_dist[k]   for k in SEVERITY_ORDER if k in sev_dist]
            sev_colors = [SEVERITY_COLORS[k] for k in SEVERITY_ORDER if k in sev_dist]
            fig_pie = go.Figure(go.Pie(
                labels=sev_labels, values=sev_values,
                marker=dict(colors=sev_colors, line=dict(color="#0f172a", width=2)),
                hole=0.45, textinfo="percent+label", textfont=dict(size=13),
                hovertemplate="<b>%{label}</b><br>Transacciones: %{value}<br>%{percent}<extra></extra>",
            ))
            fig_pie.update_layout(**PLOTLY_LAYOUT, showlegend=True, height=340,
                legend=dict(orientation="h", y=-0.08, x=0.5, xanchor="center"),
                annotations=[dict(text=f"<b>{m['n_transacciones']}</b><br>total",
                    x=0.5, y=0.5, font_size=14, showarrow=False,
                    font=dict(color="#e2e8f0"))])
            st.plotly_chart(fig_pie, use_container_width=True)

        with col_line:
            st.markdown("### Historial de Entrenamiento (Loss)")
            fig_loss = go.Figure()
            fig_loss.add_trace(go.Scatter(
                x=df_history["epoch"], y=df_history["loss"],
                mode="lines", name="Loss Total",
                line=dict(color="#818cf8", width=2.5),
                hovertemplate="Época %{x}<br>Loss: %{y:.5f}<extra></extra>"))
            fig_loss.add_trace(go.Scatter(
                x=df_history["epoch"], y=df_history["reconstruction_mse"],
                mode="lines", name="MSE Reconstrucción",
                line=dict(color="#22c55e", width=2, dash="dot"),
                hovertemplate="Época %{x}<br>MSE: %{y:.5f}<extra></extra>"))
            fig_loss.add_trace(go.Scatter(
                x=df_history["epoch"],
                y=df_history["kl_loss"] / df_history["kl_loss"].max(),
                mode="lines", name="KL Loss (norm.)",
                line=dict(color="#f97316", width=2, dash="dash"),
                hovertemplate="Época %{x}<br>KL norm: %{y:.4f}<extra></extra>"))
            fig_loss.update_layout(**PLOTLY_LAYOUT, height=340,
                xaxis=dict(title="Época", gridcolor="#334155", zeroline=False),
                yaxis=dict(title="Pérdida", gridcolor="#334155", zeroline=False),
                legend=dict(orientation="h", y=1.12, x=0, bgcolor="rgba(0,0,0,0)"))
            st.plotly_chart(fig_loss, use_container_width=True)


        # Barras: comparación splits
        st.markdown("### Comparación de Severidad: Validación vs Prueba")
        splits_data = []
        for split_key, split_label in [("validacion", "Validación"), ("prueba", "Prueba")]:
            dist = metricas[split_key]["distribucion_severidad"]
            for sev in SEVERITY_ORDER:
                splits_data.append({
                    "Split": split_label,
                    "Severidad": SEV_LABELS[sev],
                    "Transacciones": dist.get(sev, 0),
                })
        df_splits = pd.DataFrame(splits_data)
        fig_bar = px.bar(
            df_splits, x="Severidad", y="Transacciones",
            color="Severidad", barmode="group", facet_col="Split",
            color_discrete_map={SEV_LABELS[k]: SEVERITY_COLORS[k] for k in SEVERITY_ORDER},
            text="Transacciones",
        )
        fig_bar.update_traces(textposition="outside", textfont=dict(size=11),
                              marker_line_color="#0f172a", marker_line_width=1.5)
        fig_bar.update_layout(**PLOTLY_LAYOUT, showlegend=False, height=340,
            xaxis=dict(gridcolor="#334155"), yaxis=dict(gridcolor="#334155"),
            xaxis2=dict(gridcolor="#334155"), yaxis2=dict(gridcolor="#334155"))
        fig_bar.for_each_annotation(lambda a: a.update(
            text=a.text.replace("Split=", ""), font=dict(size=13, color="#cbd5e1")))
        st.plotly_chart(fig_bar, use_container_width=True)

        # Matriz de confusión
        st.markdown("### Matriz de Confusión (Split: Prueba)")
        mc = m["matriz_confusion"]
        z  = [[mc["verdaderos_negativos"], mc["falsos_positivos"]],
              [mc["falsos_negativos"],     mc["verdaderos_positivos"]]]
        fig_cm = go.Figure(go.Heatmap(
            z=z,
            x=["Predicho: Normal", "Predicho: Anomalía"],
            y=["Real: Normal", "Real: Anomalía"],
            colorscale=[[0,"#1e293b"],[0.5,"#4f46e5"],[1,"#22c55e"]],
            showscale=False,
            text=[[str(v) for v in row] for row in z],
            texttemplate="%{text}", textfont=dict(size=22, color="#f1f5f9"),
            hovertemplate="%{y} / %{x}: %{z}<extra></extra>",
        ))
        fig_cm.update_layout(**PLOTLY_LAYOUT, height=260,
                             xaxis=dict(side="top"), yaxis=dict(autorange="reversed"))
        st.plotly_chart(fig_cm, use_container_width=True)


    # ────────────────────────────────────────────────────────────────────────
    # SECCIÓN 2 — AUDITORÍA POR LOTES
    # ────────────────────────────────────────────────────────────────────────
    elif seccion == "📁 Auditoría por Lotes (CSV)":
        st.markdown("## 📁 Auditoría por Lotes")
        st.markdown(
            "Carga un CSV con transacciones evaluadas o usa el dataset incluido."
        )
        use_default = st.checkbox("Usar dataset de evaluación incluido", value=True)
        if use_default:
            df_audit = df_tx.copy()
            st.info(f"Cargadas **{len(df_audit):,}** transacciones desde `reports/evaluacion_transacciones.csv`")
        else:
            uploaded = st.file_uploader("Sube tu archivo CSV", type=["csv"])
            if uploaded is None:
                st.warning("Sube un archivo CSV para continuar.")
                st.stop()
            df_audit = pd.read_csv(uploaded)
            st.success(f"Archivo cargado: **{len(df_audit):,}** filas")

        df_audit, performance = process_batch(df_audit)

        st.markdown("---")
        fcol1, fcol2, fcol3, fcol4 = st.columns([1, 1, 1, 1])
        with fcol1:
            splits_disp = ["Todos"] + sorted(df_audit["split"].dropna().unique().tolist())
            filtro_split = st.selectbox("Filtrar por Split", splits_disp)
        with fcol2:
            sevs_disp = ["Todas"] + [SEV_LABELS[s] for s in SEVERITY_ORDER
                                     if s in df_audit["severidad"].values]
            filtro_sev = st.selectbox("Filtrar por Severidad", sevs_disp)
        with fcol3:
            busqueda_query = st.text_input("🔍 Buscar ID o Tipo", placeholder="Ej: 10042 o descuento")
        with fcol4:
            solo_anomalias = st.toggle("Solo anomalías detectadas", value=False)

        # 1. Filtrado base (Split, Rango de Monto, Búsqueda, Solo Anomalías)
        df_base = df_audit.copy()
        if filtro_split != "Todos":
            df_base = df_base[df_base["split"] == filtro_split]

        if "monto_final" in df_base.columns:
            m_min = float(df_audit["monto_final"].min())
            m_max = float(df_audit["monto_final"].max())
            rango_monto = st.slider(
                "Rango de Monto ($)",
                min_value=0.0,
                max_value=max(m_max, 10.0),
                value=(0.0, max(m_max, 10.0)),
                step=1.0,
            )
            df_base = df_base[
                (df_base["monto_final"] >= rango_monto[0]) &
                (df_base["monto_final"] <= rango_monto[1])
            ]

        if solo_anomalias:
            df_base = df_base[df_base["prediccion_anomalia"] == 1]
        if busqueda_query.strip():
            q = busqueda_query.strip().lower()
            mask_busqueda = df_base["id_transaccion"].astype(str).str.lower().str.contains(q)
            if "tipo_anomalia" in df_base.columns:
                mask_busqueda |= df_base["tipo_anomalia"].astype(str).str.lower().str.contains(q)
            df_base = df_base[mask_busqueda]

        total_transacciones_base = len(df_base)

        # 2. Filtrado específico por Severidad sobre df_base
        df_filtered = df_base.copy()
        if filtro_sev != "Todas":
            sev_key = {v: k for k, v in SEV_LABELS.items()}[filtro_sev]
            df_filtered = df_filtered[df_filtered["severidad"] == sev_key]

        n_coincidencias = len(df_filtered)
        pct_del_total = (n_coincidencias / total_transacciones_base * 100) if total_transacciones_base > 0 else 0.0

        # Anomalías en la selección para el monto en riesgo
        anomalias_lote    = df_filtered[df_filtered["prediccion_anomalia"] == 1]
        n_anomalias       = len(anomalias_lote)
        monto_riesgo_lote = anomalias_lote["monto_final"].sum() if "monto_final" in anomalias_lote.columns else 0.0

        if filtro_sev == "Todas":
            pct_anomalias_total = (n_anomalias / total_transacciones_base * 100) if total_transacciones_base > 0 else 0.0
            st.markdown(f"**{total_transacciones_base:,}** transacciones en el lote — **{n_anomalias:,}** anomalías detectadas ({pct_anomalias_total:.1f}%).")
        else:
            st.markdown(f"**{n_coincidencias:,}** transacciones con severidad **{filtro_sev}** ({pct_del_total:.1f}% del total de **{total_transacciones_base:,}** transacciones).")

        bk1, bk2, bk3 = st.columns(3)
        with bk1:
            st.metric("Total Transacciones", f"{total_transacciones_base:,}")
        with bk2:
            if filtro_sev == "Todas":
                pct_anom = (n_anomalias / total_transacciones_base * 100) if total_transacciones_base > 0 else 0.0
                st.metric("Anomalías Detectadas", f"{n_anomalias:,}",
                          delta=f"{pct_anom:.1f}% del total", delta_color="inverse")
            else:
                color_delta = "normal" if filtro_sev == "Normal" else "inverse"
                st.metric(f"Transacciones ({filtro_sev})", f"{n_coincidencias:,}",
                          delta=f"{pct_del_total:.1f}% del total", delta_color=color_delta)
        with bk3:
            st.metric("Monto en Riesgo", f"${monto_riesgo_lote:,.2f}")

        gc1, gc2 = st.columns(2)
        with gc1:
            st.markdown("#### Distribución del Error de Reconstrucción")
            fig_hist = px.histogram(
                df_filtered, x="reconstruction_error", color="severidad",
                color_discrete_map=SEVERITY_COLORS,
                category_orders={"severidad": SEVERITY_ORDER},
                nbins=60, barmode="overlay", opacity=0.75,
                labels={"reconstruction_error": "Error de Reconstrucción", "severidad": "Severidad"},
            )

            # Configurar qué líneas de umbral y qué rango del eje X mostrar según el filtro de severidad
            if filtro_sev == "Normal":
                keys_to_show = ["umbral_baja"]
                x_min = 0.0
                x_max = max(df_filtered["reconstruction_error"].max() * 1.08, umbrales["umbral_baja"] * 1.15) if len(df_filtered) else 0.1
            elif filtro_sev == "Baja":
                keys_to_show = ["umbral_baja", "umbral_media"]
                x_min = min(df_filtered["reconstruction_error"].min(), umbrales["umbral_baja"]) * 0.96 if len(df_filtered) else 0.0
                x_max = max(df_filtered["reconstruction_error"].max(), umbrales["umbral_media"]) * 1.04 if len(df_filtered) else 0.15
            elif filtro_sev == "Media":
                keys_to_show = ["umbral_media", "umbral_alta"]
                x_min = min(df_filtered["reconstruction_error"].min(), umbrales["umbral_media"]) * 0.96 if len(df_filtered) else 0.05
                x_max = max(df_filtered["reconstruction_error"].max(), umbrales["umbral_alta"]) * 1.04 if len(df_filtered) else 0.25
            elif filtro_sev == "Alta":
                keys_to_show = ["umbral_alta"]
                x_min = umbrales["umbral_alta"] * 0.96
                x_max = float(df_filtered["reconstruction_error"].quantile(0.99) * 1.15) if len(df_filtered) else umbrales["umbral_alta"] * 2.0
            else:  # "Todas"
                keys_to_show = ["umbral_baja", "umbral_media", "umbral_alta"]
                x_min = 0.0
                # Enfocar en el rango donde está la mayoría de datos sin comprimir todo
                x_max = float(umbrales.get("umbral_alta", 0.5) * 2.5) if len(df_filtered) else 0.3

            # Desplazamientos visuales de anotación para evitar colisiones
            thresh_meta = {
                "umbral_baja":  ("Baja",  "top left", 0),
                "umbral_media": ("Media", "top right", -15),
                "umbral_alta":  ("Alta",  "top right", 25),
            }

            for key in keys_to_show:
                if key in umbrales and key in thresh_meta:
                    label, pos, y_shift = thresh_meta[key]
                    fig_hist.add_vline(
                        x=umbrales[key],
                        line_dash="dash",
                        line_color=SEVERITY_COLORS[key.replace("umbral_", "")],
                        annotation_text=f" <b>{label}</b>",
                        annotation_position=pos,
                        annotation_font_color=SEVERITY_COLORS[key.replace("umbral_", "")],
                        annotation_font_size=10,
                        annotation_yshift=y_shift,
                    )

            fig_hist.update_layout(**PLOTLY_LAYOUT, height=340,
                xaxis=dict(gridcolor="#334155", range=[x_min, max(x_max, x_min + 0.05)]),
                yaxis=dict(gridcolor="#334155", title="Frecuencia"),
                legend=dict(title="Severidad"))
            st.plotly_chart(fig_hist, use_container_width=True)

        with gc2:
            st.markdown("#### Monto en Riesgo por Severidad")
            risk_by_sev = (
                df_filtered[df_filtered["prediccion_anomalia"] == 1]
                .groupby("severidad")["monto_final"].sum()
                .reindex(SEVERITY_ORDER).dropna().reset_index()
            )
            risk_by_sev.columns = ["severidad", "monto"]
            risk_by_sev["label"] = risk_by_sev["severidad"].map(SEV_LABELS)
            fig_risk = px.bar(
                risk_by_sev, x="label", y="monto", color="severidad",
                color_discrete_map=SEVERITY_COLORS,
                text=risk_by_sev["monto"].apply(lambda v: f"${v:,.0f}"),
                labels={"label": "Severidad", "monto": "Monto ($)"},
            )
            fig_risk.update_traces(textposition="outside", marker_line_width=0)
            fig_risk.update_layout(**PLOTLY_LAYOUT, height=340, showlegend=False,
                xaxis=dict(gridcolor="#334155"), yaxis=dict(gridcolor="#334155"))
            st.plotly_chart(fig_risk, use_container_width=True)

        st.markdown("#### Detalle de Transacciones")
        d_col1, d_col2 = st.columns([3, 1])
        with d_col2:
            filas_opc = st.selectbox("Filas a mostrar", ["Todas (Sin límite)", 100, 500, 1000], index=0)

        display_cols = ["id_transaccion","split","monto_final","tipo_anomalia",
                        "reconstruction_error","severidad","prediccion_anomalia"]
        available_cols = [c for c in display_cols if c in df_filtered.columns]

        # Motivo legible calculado ANTES de renombrar/formatear columnas,
        # para tener acceso a los nombres originales (monto_final, tipo_anomalia, ...)
        motivo_series = df_filtered.apply(lambda r: explicar_anomalia(r, estilo="tecnico"), axis=1)

        df_display = df_filtered[available_cols].copy()
        df_display["motivo"] = motivo_series.values
        if "monto_final" in df_display.columns:
            df_display["monto_final"] = df_display["monto_final"].apply(lambda x: f"${x:,.2f}")
        if "reconstruction_error" in df_display.columns:
            df_display["reconstruction_error"] = df_display["reconstruction_error"].apply(lambda x: f"{x:.6f}")
        if "prediccion_anomalia" in df_display.columns:
            df_display["prediccion_anomalia"] = df_display["prediccion_anomalia"].map({0:"✅ Normal",1:"⚠️ Anomalía"})
        df_display = df_display.rename(columns={
            "id_transaccion":"ID","split":"Split","monto_final":"Monto",
            "tipo_anomalia":"Tipo","reconstruction_error":"Error Reconstrucción",
            "severidad":"Severidad","prediccion_anomalia":"Predicción","motivo":"Motivo"})

        if isinstance(filas_opc, int):
            df_show_table = df_display.head(filas_opc)
            st.caption(f"Mostrando {len(df_show_table):,} de {len(df_display):,} transacciones encontradas.")
        else:
            df_show_table = df_display
            st.caption(f"Mostrando **todas las {len(df_show_table):,}** transacciones encontradas sin límite.")

        st.dataframe(df_show_table, use_container_width=True, height=450)

        # ── Descarga en Excel, ordenada, coloreada por severidad y con filtros ──
        df_export = df_filtered[available_cols].copy()
        df_export["motivo"] = motivo_series.values
        if "prediccion_anomalia" in df_export.columns:
            df_export["prediccion_anomalia"] = df_export["prediccion_anomalia"].map(
                {0: "Normal", 1: "Anomalía"})
        sev_export_series = (
            df_filtered["severidad"].str.lower() if "severidad" in df_filtered.columns else None
        )
        df_export = df_export.rename(columns={
            "id_transaccion": "ID", "split": "Split", "monto_final": "Monto ($)",
            "tipo_anomalia": "Tipo", "reconstruction_error": "Error Reconstrucción",
            "severidad": "Severidad", "prediccion_anomalia": "Predicción", "motivo": "Motivo",
        })
        if "Severidad" in df_export.columns:
            df_export["Severidad"] = df_export["Severidad"].str.capitalize()

        excel_bytes = build_styled_excel_bytes([{
            "name": "Auditoría Filtrada",
            "df": df_export,
            "sev_series": sev_export_series,
            "currency_cols": ["Monto ($)"],
            "decimal_cols": ["Error Reconstrucción"],
        }])

        st.download_button(
            "⬇️ Descargar resultados filtrados (.xlsx)",
            data=excel_bytes,
            file_name="auditoria_filtrada.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


    # ────────────────────────────────────────────────────────────────────────
    # SECCIÓN 3 — EVALUACIÓN EN VIVO
    # ────────────────────────────────────────────────────────────────────────
    elif seccion == "⚡ Verificar Venta":
        st.markdown("## ⚡ Verificar Venta")
        st.markdown(
            "Ingresa los datos de una transacción. "
            "El sistema calculará el error de reconstrucción con el modelo VAE real y clasificará su severidad."
        )
        with st.form("form_live"):
            st.markdown("### Datos de la Transacción")
            lc1, lc2, lc3 = st.columns(3)
            with lc1:
                monto     = st.number_input("Monto ($)", min_value=0.0, max_value=5000.0,
                                            value=15.50, step=0.5, format="%.2f")
                descuento = st.number_input("Descuento aplicado ($)", min_value=0.0,
                                            max_value=500.0, value=0.0, step=0.5)
            with lc2:
                hora       = st.slider("Hora del día", 0, 23, 12)
                dia_semana = st.selectbox("Día de la semana",
                    ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"])
            with lc3:
                metodo_pago = st.selectbox("Método de pago",
                    ["Efectivo","Tarjeta crédito","Tarjeta débito","Transferencia","QR / Billetera digital"])
                num_items = st.number_input("Número de ítems", min_value=1, max_value=50, value=3, step=1)
            submitted = st.form_submit_button("🔍 Evaluar Venta", use_container_width=True)

        if submitted:
            tx_data = {
                "monto": monto,
                "descuento": descuento,
                "hora": hora,
                "dia_semana": dia_semana,
                "metodo_pago": metodo_pago,
                "num_items": num_items,
            }
            resultado, performance = evaluate_raw_transaction(tx_data)

            rec_error = resultado["reconstruction_error"]
            severidad = resultado["severidad"]
            es_anomalia = resultado["es_anomalia"]

            st.markdown("---")
            st.markdown("### Resultado de la Evaluación")
            sev_color = SEVERITY_COLORS[severidad]
            icono     = "🚨" if severidad == "alta" else ("⚠️" if severidad in ("media","baja") else "✅")

            res1, res2 = st.columns([1, 2])
            with res1:
                st.markdown(
                    f"""<div style="background:{sev_color}18; border:2px solid {sev_color};
                        border-radius:16px; padding:28px; text-align:center;">
                      <div style="font-size:3rem;">{icono}</div>
                      <div style="font-size:1.6rem; font-weight:700; color:{sev_color};
                                  margin-top:8px;">{SEV_LABELS[severidad]}</div>
                      <div style="font-size:0.9rem; color:#94a3b8; margin-top:6px;">
                        {"⚠️ Anomalía detectada" if es_anomalia else "✅ Transacción normal"}
                      </div>
                      <hr style="border-color:#334155; margin:14px 0;">
                      <div style="font-size:0.85rem; color:#94a3b8;">Error de reconstrucción</div>
                      <div style="font-size:1.5rem; font-weight:600; color:#e2e8f0;">
                        {rec_error:.6f}
                      </div>
                    </div>""", unsafe_allow_html=True)
            with res2:
                fig_gauge = go.Figure(go.Indicator(
                    mode="gauge+number",
                    value=rec_error,
                    number=dict(valueformat=".5f", font=dict(color="#e2e8f0", size=20)),
                    gauge=dict(
                        axis=dict(range=[0, max(umbrales["umbral_alta"]*2, rec_error*1.2)],
                                  tickcolor="#94a3b8", tickfont=dict(color="#94a3b8", size=10)),
                        bar=dict(color=sev_color, thickness=0.3),
                        bgcolor="#1e293b", borderwidth=0,
                        steps=[
                            dict(range=[0, umbrales["umbral_baja"]], color="rgba(34,197,94,0.13)"),
                            dict(range=[umbrales["umbral_baja"], umbrales["umbral_media"]], color="rgba(234,179,8,0.13)"),
                            dict(range=[umbrales["umbral_media"], umbrales["umbral_alta"]], color="rgba(249,115,22,0.13)"),
                            dict(range=[umbrales["umbral_alta"], umbrales["umbral_alta"]*3], color="rgba(239,68,68,0.13)"),
                        ],
                        threshold=dict(line=dict(color=sev_color, width=3),
                                       thickness=0.8, value=rec_error),
                    ),
                    title=dict(text="Error de Reconstrucción VAE",
                               font=dict(color="#94a3b8", size=13)),
                ))
                fig_gauge.update_layout(**PLOTLY_LAYOUT, height=260)
                st.plotly_chart(fig_gauge, use_container_width=True)

            det1, det2 = st.columns(2)
            with det1:
                st.markdown(f"""| Campo | Valor |
|---|---|
| Monto | **${monto:,.2f}** |
| Descuento | **${descuento:,.2f}** |
| Ratio descuento | **{descuento/max(monto,1)*100:.1f}%** |
| N° ítems | **{num_items}** |""")
            with det2:
                st.markdown(f"""| Campo | Valor |
|---|---|
| Hora | **{hora:02d}:00** |
| Día | **{dia_semana}** |
| Método de pago | **{metodo_pago}** |
| Horario inusual | **{"Sí ⚠️" if hora < 6 or hora > 22 else "No ✅"}** |""")

            if es_anomalia:
                st.markdown("#### Factores que contribuyen a la anomalía")
                motivo_modelo = resultado.get("motivo_tecnico")
                if motivo_modelo:
                    st.warning(
                        f"📊 Según el error de reconstrucción del VAE, lo que más se "
                        f"desvió de lo normal fue: **{motivo_modelo}**"
                    )
                else:
                    # Respaldo: el error está repartido entre varias variables sin
                    # que ninguna concentre una parte clara — se listan indicios
                    # adicionales calculados a mano, sin inventar una causa única.
                    factores = []
                    if monto > 200:
                        factores.append(f"💰 Monto elevado (${monto:.2f}) supera el umbral típico de $200")
                    if monto < 1.0:
                        factores.append(f"💸 Monto muy bajo (${monto:.2f}) — posible error o transacción de prueba")
                    if descuento > 0 and (descuento / max(monto,1)) > 0.3:
                        factores.append(f"🎟️ Descuento del {descuento/max(monto,1)*100:.0f}% supera el 30% permitido")
                    if hora < 6 or hora > 22:
                        factores.append(f"🕐 Hora inusual ({hora:02d}:00) fuera del horario de operación normal")
                    if num_items > 20:
                        factores.append(f"📦 Número de ítems ({num_items}) inusualmente alto")
                    if not factores:
                        factores.append("📊 Patrón general inusual detectado por el VAE, repartido entre varias variables")
                    for f_msg in factores:
                        st.warning(f_msg)
            else:
                st.success("La transacción está dentro de los patrones normales. No se detectaron anomalías.")
        else:
            st.info("Completa el formulario y presiona **Evaluar Transacción** para obtener el diagnóstico.")
            uc1, uc2, uc3, uc4 = st.columns(4)
            for col, sev, rng in zip([uc1,uc2,uc3,uc4], SEVERITY_ORDER, [
                f"< {umbrales['umbral_baja']:.5f}",
                f"{umbrales['umbral_baja']:.5f} – {umbrales['umbral_media']:.5f}",
                f"{umbrales['umbral_media']:.5f} – {umbrales['umbral_alta']:.5f}",
                f"≥ {umbrales['umbral_alta']:.5f}",
            ]):
                with col:
                    st.markdown(
                        f"""<div class="kpi-card" style="border-color:{SEVERITY_COLORS[sev]}44;">
                          <div style="font-size:1.1rem;font-weight:700;color:{SEVERITY_COLORS[sev]};">
                            {SEV_LABELS[sev]}</div>
                          <div style="font-size:0.8rem;color:#94a3b8;margin-top:8px;">{rng}</div>
                        </div>""", unsafe_allow_html=True)



    # ────────────────────────────────────────────────────────────────────────
    # SECCIÓN 4 — LOGS Y RENDIMIENTO DEL SISTEMA
    # ────────────────────────────────────────────────────────────────────────
    elif seccion == "📜 Logs y Rendimiento Sistema":
        st.markdown("## 📜 Logs y Rendimiento del Sistema")
        st.markdown("Monitoreo de latencia, throughput, auditoría de eventos y pruebas de concurrencia multiusuario.")

        l1, l2, l3 = st.columns(3)
        with l1:
            st.markdown(
                """<div class="kpi-card">
                  <div class="kpi-title">Inferencia y Modelo</div>
                  <div class="kpi-value" style="color:#38bdf8;">PyTorch VAE</div>
                  <div class="kpi-sub">Preprocesador: Scikit-Learn Joblib</div>
                </div>""", unsafe_allow_html=True)
        with l2:
            st.markdown(
                """<div class="kpi-card">
                  <div class="kpi-title">Soporte Concurrente</div>
                  <div class="kpi-value" style="color:#22c55e;">3+ Usuarios</div>
                  <div class="kpi-sub">Thread-Safe & Async FastAPI</div>
                </div>""", unsafe_allow_html=True)
        with l3:
            st.markdown(
                """<div class="kpi-card">
                  <div class="kpi-title">Archivo de Logs</div>
                  <div class="kpi-value" style="color:#a78bfa;">app.log</div>
                  <div class="kpi-sub">Ubicación: logs/app.log</div>
                </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        st.markdown("### 🔍 Logs en Tiempo Real")
        if LOG_FILE.exists():
            with open(LOG_FILE, "r", encoding="utf-8") as log_f:
                logs_lines = [line.strip() for line in log_f.readlines() if line.strip()]

            num_logs = st.slider("Número de líneas a mostrar", 10, 200, 50)
            visible_lines = logs_lines[-num_logs:]
            logs_to_show = "\n".join(visible_lines)
            st.text_area("Eventos del sistema (logs/app.log):", value=logs_to_show, height=320)
            st.caption(f"Mostrando las últimas {min(num_logs, len(logs_lines))} de {len(logs_lines)} entradas de log.")

            # ── Resumen calculado en vivo de las líneas visibles ────────────────
            # Solo se agregan eventos "REAL_*": los que sí pasan por el modelo VAE
            # real (src/audit_service.py). El log también conserva entradas
            # antiguas de antes de conectar el backend real, con cifras que ya
            # no son representativas — se excluyen de las estadísticas.
            df_logs = parse_log_lines(visible_lines)
            df_real = df_logs[df_logs["event"].str.startswith("REAL_")]
            df_perf = df_real.dropna(subset=["throughput"])

            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("### 📊 Resumen de Rendimiento (calculado de los logs reales)")
            st.caption(
                "Solo incluye eventos `REAL_*`, generados por el modelo VAE real "
                "en producción (excluye entradas de logs previas a la integración "
                "del backend)."
            )

            lk1, lk2, lk3, lk4 = st.columns(4)
            lk1.metric("Eventos analizados", f"{len(df_logs):,}")
            lk2.metric("Eventos reales (VAE)", f"{len(df_real):,}")
            lk3.metric(
                "Throughput promedio",
                f"{df_perf['throughput'].mean():,.0f} trans/s" if len(df_perf) else "—",
            )
            lk4.metric(
                "Latencia promedio",
                f"{df_real['latencia_ms'].dropna().mean():.3f} ms"
                if df_real["latencia_ms"].notna().any() else "—",
            )

            if len(df_perf) >= 2:
                fig_perf = go.Figure()
                fig_perf.add_trace(go.Scatter(
                    y=df_perf["throughput"], mode="lines+markers",
                    line=dict(color=SEVERITY_COLORS["normal"], width=2),
                    marker=dict(size=5),
                    name="Throughput (trans/s)",
                ))
                perf_layout = {**PLOTLY_LAYOUT, "margin": dict(l=10, r=10, t=10, b=10)}
                fig_perf.update_layout(
                    **perf_layout,
                    height=280,
                    xaxis_title="Evento (orden cronológico)",
                    yaxis_title="Throughput (trans/s)",
                )
                st.plotly_chart(fig_perf, use_container_width=True)
            else:
                st.caption(
                    "Aún no hay suficientes eventos con throughput en la ventana "
                    "visible para graficar una tendencia. Sube el número de líneas "
                    "a mostrar o genera más tráfico (Auditoría por Lotes / Verificar Venta)."
                )
        else:
            st.info("El archivo `logs/app.log` aún no contiene entradas.")


# ════════════════════════════════════════════════════════════════════════════
# ROL NEGOCIO — RESTAURANTE ROSITA
# ════════════════════════════════════════════════════════════════════════════
elif st.session_state.rol == "Negocio":

    # ── Bloque de bienvenida corporativo ─────────────────────────────────────
    st.markdown("""
    <div class="rosita-welcome">
      <div style="display:flex; align-items:center; gap:16px; flex-wrap:wrap;">
        <div style="font-size:3rem;">🍽️</div>
        <div>
          <h1 style="color:#f97316; margin:0; font-size:1.8rem;">
            ZeroAnomalías — Restaurante Rosita
          </h1>
          <p style="color:#94a3b8; margin:4px 0 0; font-size:0.95rem;">
            Bienvenida, Rosita. Sube tu reporte de ventas y el sistema identificará
            automáticamente las transacciones que requieren tu atención.
          </p>
        </div>
      </div>
      <div style="display:flex; gap:32px; margin-top:20px; flex-wrap:wrap;">
        <div style="text-align:center;">
          <div style="font-size:1.4rem;">⚡</div>
          <div style="font-size:0.8rem; color:#64748b;">Análisis instantáneo</div>
        </div>
        <div style="text-align:center;">
          <div style="font-size:1.4rem;">🔒</div>
          <div style="font-size:0.8rem; color:#64748b;">Datos privados y seguros</div>
        </div>
        <div style="text-align:center;">
          <div style="font-size:1.4rem;">📊</div>
          <div style="font-size:0.8rem; color:#64748b;">Reportes descargables</div>
        </div>
        <div style="text-align:center;">
          <div style="font-size:1.4rem;">🎯</div>
          <div style="font-size:0.8rem; color:#64748b;">Alertas de riesgo claras</div>
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Uploader de Excel ─────────────────────────────────────────────────────
    st.markdown("### 📂 Cargar Reporte de Ventas")

    st.info(
        "**Formato requerido:** El archivo Excel debe contener las siguientes columnas: "
        "`id_transaccion`, `fecha_hora`, `cajero`, `mesa`, `monto`, "
        "`descuento_pct`, `metodo_pago`, `tipo_transaccion`"
    )

    excel_file = st.file_uploader(
        "Arrastra aquí tu archivo de ventas o haz clic para seleccionarlo",
        type=["xlsx", "xls"],
        help="Solo se aceptan archivos Excel (.xlsx o .xls)",
    )


    # ── Procesamiento tras subir el Excel ─────────────────────────────────────
    if excel_file is not None:
        try:
            df_rosita = pd.read_excel(excel_file)
        except Exception as e:  # pylint: disable=broad-exception-caught
            # Intencional: pd.read_excel puede fallar por muchas causas distintas
            # (ImportError de openpyxl, archivo corrupto, formato inválido, etc.)
            # y queremos un único mensaje amigable para la usuaria de negocio.
            st.error(f"No se pudo leer el archivo: {e}")
            st.stop()

        # Validar columnas requeridas
        REQUIRED_COLS = ["id_transaccion","fecha_hora","cajero","mesa",
                         "monto","descuento_pct","metodo_pago","tipo_transaccion"]
        missing = [c for c in REQUIRED_COLS if c not in df_rosita.columns]
        if missing:
            st.error(
                f"El archivo no tiene las columnas requeridas: **{', '.join(missing)}**\n\n"
                "Revisa que tu Excel tenga exactamente los encabezados indicados."
            )
            st.stop()

        # ── Procesamiento VAE PyTorch Real ───────────────────────────────────
        with st.spinner("Analizando transacciones con el modelo VAE de PyTorch real..."):
            df_proc, performance_rosita = process_raw_batch(df_rosita)
            # OJO: process_raw_batch reordena internamente por error de
            # reconstrucción (mayor a menor) y resetea el índice. Copiar las
            # columnas de vuelta por posición mezclaba la severidad de una
            # transacción con los datos de otra. Se une por id_transaccion,
            # que es estable sin importar el orden interno.
            df_proc_por_id = df_proc.set_index("id_transaccion")
            df_rosita["_error_vae"] = df_rosita["id_transaccion"].map(df_proc_por_id["reconstruction_error"])
            df_rosita["_severidad"] = df_rosita["id_transaccion"].map(df_proc_por_id["severidad"])
            df_rosita["motivo_tecnico"] = df_rosita["id_transaccion"].map(df_proc_por_id["motivo_tecnico"])
            df_rosita["motivo_negocio"] = df_rosita["id_transaccion"].map(df_proc_por_id["motivo_negocio"])
            # "baja" se agrupa visualmente con "normal" en el perfil de Negocio
            # (ver simplify_sev más abajo). _es_anomalia usa esa MISMA definición
            # simplificada para que el KPI de arriba y el gráfico de abajo coincidan.
            df_rosita["_es_anomalia"] = df_rosita["_severidad"].isin(["media", "alta"])
            n = len(df_rosita)

        st.success(f"✅ Análisis completado — **{n:,}** transacciones procesadas.")
        st.markdown("---")

        # ── KPIs ──────────────────────────────────────────────────────────────
        st.markdown("### 📈 Resumen Ejecutivo")

        # Transacciones con alerta real (media o alta) — "baja" cuenta como normal aquí
        anomalias_r = df_rosita[df_rosita["_es_anomalia"]].copy()

        # "Posibles pérdidas" = solo transacciones comprometidas: severidad media o alta
        comprometidas_r = df_rosita[df_rosita["_severidad"].isin(["media", "alta"])].copy()

        # Columna de monto: acepta "monto" o "monto_final"
        col_monto = "monto" if "monto" in df_rosita.columns else (
                    "monto_final" if "monto_final" in df_rosita.columns else None)

        if col_monto:
            monto_riesgo_r = pd.to_numeric(
                comprometidas_r[col_monto], errors="coerce"
            ).fillna(0).sum()
        else:
            monto_riesgo_r = 0

        total_tx    = len(df_rosita)
        n_anomalias = len(anomalias_r)
        pct_riesgo  = (n_anomalias / total_tx * 100) if total_tx else 0

        rk1, rk2, rk3 = st.columns(3)
        with rk1:
            st.markdown(
                f"""<div class="kpi-card">
                  <div class="kpi-title">Total de Transacciones Revisadas</div>
                  <div class="kpi-value">{total_tx:,}</div>
                  <div class="kpi-sub">en este reporte</div>
                </div>""", unsafe_allow_html=True)
        with rk2:
            st.markdown(
                f"""<div class="kpi-card">
                  <div class="kpi-title">Transacciones con Alerta</div>
                  <div class="kpi-value" style="color:#f97316;">{n_anomalias:,}</div>
                  <div class="kpi-sub">{pct_riesgo:.1f}% del total</div>
                </div>""", unsafe_allow_html=True)
        with rk3:
            st.markdown(
                f"""<div class="kpi-card">
                  <div class="kpi-title">Posibles Pérdidas (Monto en Riesgo)</div>
                  <div class="kpi-value" style="color:#ef4444;">${monto_riesgo_r:,.2f}</div>
                  <div class="kpi-sub">solo alertas media y alta</div>
                </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)


        # ── Gráfico de distribución de severidad ──────────────────────────────
        st.markdown("### 🔎 Distribución de Transacciones por Estado")

        # Solo mostramos Normal, Media y Alta (sin "baja") al perfil de negocio
        SEV_NEGOCIO   = ["normal", "media", "alta"]
        SEV_NEG_LABEL = {"normal": "Normal ✅", "media": "Alerta Media ⚠️", "alta": "Alerta Alta 🚨"}
        SEV_NEG_COLOR = {"normal": "#22c55e", "media": "#f97316", "alta": "#ef4444"}

        # Agrupamos "baja" dentro de "normal" para simplificar la vista de negocio
        def simplify_sev(s):
            """Agrupa la severidad 'baja' dentro de 'normal' para la vista de Negocio."""
            return s if s in ("normal", "media", "alta") else "normal"

        df_rosita["_sev_simple"] = df_rosita["_severidad"].apply(simplify_sev)

        sev_counts = (
            df_rosita["_sev_simple"]
            .value_counts()
            .reindex(SEV_NEGOCIO, fill_value=0)
            .reset_index()
        )
        sev_counts.columns = ["severidad", "cantidad"]
        sev_counts["label"] = sev_counts["severidad"].map(SEV_NEG_LABEL)
        sev_counts["color"] = sev_counts["severidad"].map(SEV_NEG_COLOR)

        gc1_r, gc2_r = st.columns([1, 1.4])
        with gc1_r:
            fig_pie_r = go.Figure(go.Pie(
                labels=sev_counts["label"],
                values=sev_counts["cantidad"],
                marker=dict(
                    colors=sev_counts["color"].tolist(),
                    line=dict(color="#0f172a", width=2)
                ),
                hole=0.5,
                textinfo="percent+label",
                textfont=dict(size=12),
                hovertemplate="<b>%{label}</b><br>%{value} transacciones<br>%{percent}<extra></extra>",
            ))
            fig_pie_r.update_layout(
                **PLOTLY_LAYOUT, height=320, showlegend=False,
                annotations=[dict(
                    text=f"<b>{total_tx}</b><br>total",
                    x=0.5, y=0.5, font_size=13, showarrow=False,
                    font=dict(color="#e2e8f0"),
                )])
            st.plotly_chart(fig_pie_r, use_container_width=True)

        with gc2_r:
            fig_bar_r = px.bar(
                sev_counts, x="label", y="cantidad",
                color="severidad",
                color_discrete_map=SEV_NEG_COLOR,
                text="cantidad",
                labels={"label": "Estado", "cantidad": "N° Transacciones"},
            )
            fig_bar_r.update_traces(
                textposition="outside", marker_line_width=0,
                textfont=dict(size=13))
            fig_bar_r.update_layout(
                **PLOTLY_LAYOUT, showlegend=False, height=320,
                xaxis=dict(gridcolor="#334155"),
                yaxis=dict(gridcolor="#334155"))
            st.plotly_chart(fig_bar_r, use_container_width=True)


        # ── Tabla de resultados con estilos condicionales ─────────────────────
        st.markdown("### 📋 Tabla de Resultados")
        st.caption("Las transacciones están ordenadas por nivel de alerta (Altas primero).")

        r_fil1, r_fil2 = st.columns([2, 1])
        with r_fil1:
            busqueda_rosita = st.text_input("🔍 Buscar por Cajero, Mesa, Tipo o ID", placeholder="Ej: Cajero_1 o Mesa_3")
        with r_fil2:
            filas_rosita_opc = st.selectbox("Filas a mostrar (Rosita)", ["Todas (Sin límite)", 500, 1000], index=0)

        # Orden de severidad para sort
        SEV_SORT = {"alta": 0, "media": 1, "baja": 2, "normal": 3}
        df_rosita["_sev_order"] = df_rosita["_severidad"].map(SEV_SORT)

        # Seleccionar columnas de negocio (originales + estado)
        cols_negocio = [c for c in REQUIRED_COLS if c in df_rosita.columns]
        df_result = (
            df_rosita[cols_negocio + ["_sev_simple", "_sev_order", "motivo_tecnico", "motivo_negocio"]]
            .sort_values("_sev_order")
            .drop(columns=["_sev_order"])
            .rename(columns={"_sev_simple": "estado"})
            .reset_index(drop=True)
        )

        if busqueda_rosita.strip():
            q_r = busqueda_rosita.strip().lower()
            mask = pd.Series(False, index=df_result.index)
            for c in df_result.columns:
                mask |= df_result[c].astype(str).str.lower().str.contains(q_r)
            df_result = df_result[mask]

        # df_result se queda con los valores originales (estado en minúscula,
        # monto numérico) para el Excel. df_screen es una copia solo para
        # mostrar en pantalla, con emojis y monto formateado como texto.
        df_screen = df_result.copy()

        # Motivo en lenguaje simple (calculado ANTES de formatear monto/estado
        # para pantalla, usando los valores numéricos reales)
        motivo_rosita = df_result.apply(lambda r: explicar_anomalia(r, estilo="negocio"), axis=1)
        df_result = df_result.drop(columns=["motivo_tecnico", "motivo_negocio"])
        df_screen = df_screen.drop(columns=["motivo_tecnico", "motivo_negocio"])
        df_result["motivo"] = motivo_rosita.values
        df_screen["motivo"] = motivo_rosita.values

        # Mapear etiquetas amigables (solo en pantalla)
        df_screen["estado"] = df_screen["estado"].map(
            {"normal": "✅ Normal", "media": "⚠️ Alerta Media", "alta": "🚨 Alerta Alta"})

        # Formatear monto (solo en pantalla)
        if "monto" in df_screen.columns:
            df_screen["monto"] = df_screen["monto"].apply(
                lambda x: f"${pd.to_numeric(x, errors='coerce'):,.2f}"
                if pd.notna(pd.to_numeric(x, errors='coerce')) else x)

        # Función de estilos por fila según estado
        def style_row(row):
            """Devuelve el color de fondo de la fila según su estado de severidad."""
            estado = str(row.get("estado", ""))
            if "Alerta Alta" in estado:
                return ["background-color:#ef444420; color:#fca5a5"] * len(row)
            if "Alerta Media" in estado:
                return ["background-color:#f9731620; color:#fdba74"] * len(row)
            return ["background-color:#22c55e15; color:#86efac"] * len(row)

        if isinstance(filas_rosita_opc, int):
            df_show = df_screen.head(filas_rosita_opc)
            st.caption(f"Mostrando {len(df_show):,} de {len(df_screen):,} transacciones.")
        else:
            df_show = df_screen
            st.caption(f"Mostrando **todas las {len(df_show):,}** transacciones procesadas sin límite.")

        styled_df = df_show.style.apply(style_row, axis=1)
        st.dataframe(styled_df, use_container_width=True, height=450, hide_index=True)

        # ── Botón de descarga Excel: ordenado, coloreado por severidad y con filtros ──
        st.markdown("---")
        st.markdown("### ⬇️ Descargar Reporte Procesado")

        COLS_RENAME_ROSITA = {
            "id_transaccion": "ID Transacción", "fecha_hora": "Fecha y Hora",
            "cajero": "Cajero", "mesa": "Mesa", "monto": "Monto ($)",
            "descuento_pct": "Descuento (%)", "metodo_pago": "Método de Pago",
            "tipo_transaccion": "Tipo de Transacción", "estado": "Severidad",
            "motivo": "Motivo",
        }

        # Hoja 1: reporte completo (valores reales, no texto formateado)
        sev_completo = df_result["estado"].str.lower()
        df_excel_completo = df_result.rename(columns=COLS_RENAME_ROSITA)
        df_excel_completo["Severidad"] = df_excel_completo["Severidad"].str.capitalize()

        # Hoja 2: solo alertas (media o alta)
        df_result_alertas = df_result[df_result["estado"] != "normal"].copy()
        sev_alertas = df_result_alertas["estado"].str.lower()
        df_excel_alertas = df_result_alertas.rename(columns=COLS_RENAME_ROSITA)
        df_excel_alertas["Severidad"] = df_excel_alertas["Severidad"].str.capitalize()

        # Hoja 3: resumen ejecutivo (la fila de pérdidas se resalta en rojo)
        df_resumen = pd.DataFrame({
            "Métrica": [
                "Total transacciones revisadas",
                "Alertas detectadas",
                "% de alertas",
                "Posibles pérdidas (monto en riesgo)",
                "  → Criterio",
            ],
            "Valor": [
                total_tx,
                n_anomalias,
                f"{pct_riesgo:.1f}%",
                f"${monto_riesgo_r:,.2f}",
                "Suma de montos con severidad Media o Alta únicamente",
            ],
        })

        excel_bytes_rosita = build_styled_excel_bytes([
            {
                "name": "Reporte Completo", "df": df_excel_completo,
                "sev_series": sev_completo, "currency_cols": ["Monto ($)"],
            },
            {
                "name": "Solo Alertas", "df": df_excel_alertas,
                "sev_series": sev_alertas, "currency_cols": ["Monto ($)"],
            },
            {
                "name": "Resumen Ejecutivo", "df": df_resumen,
                "highlight_col": "Métrica", "highlight_values": ["Pérdidas"],
            },
        ])

        dcol1, dcol2 = st.columns([2, 1])
        with dcol1:
            st.download_button(
                label="⬇️  Descargar reporte completo en Excel (.xlsx)",
                data=excel_bytes_rosita,
                file_name="reporte_auditoria_rosita.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with dcol2:
            st.markdown(
                """<div style="background:#1e293b; border:1px solid #334155;
                    border-radius:8px; padding:12px 16px; font-size:0.82rem; color:#94a3b8;">
                  📄 3 hojas incluidas (coloreadas y con filtros):<br>
                  • Reporte Completo<br>
                  • Solo Alertas<br>
                  • Resumen Ejecutivo
                </div>""", unsafe_allow_html=True)

    else:
        # Estado inicial — sin archivo cargado
        st.markdown("""
        <div style="text-align:center; padding:60px 20px; color:#475569;">
          <div style="font-size:4rem; margin-bottom:16px;">📂</div>
          <h3 style="color:#64748b; font-weight:500;">
            Sube tu archivo de ventas para comenzar
          </h3>
          <p style="font-size:0.9rem; margin-top:8px;">
            El sistema analizará cada transacción y te mostrará cuáles requieren revisión.
          </p>
        </div>
        """, unsafe_allow_html=True)